from openpyxl.reader.excel import load_workbook

class Item:
    def __init__(self, name, rank):
        self.name = name
        self.rank = rank

    def __repr__(self):
        return self.name + "':" + str(self.rank)

def calculate():
    girls_workbook = load_workbook(filename="poll/FILLES.xlsx")
    boys_workbook = load_workbook(filename="poll/GARCONS.xlsx")

    girls = {}
    boys = {}
    attribution = {}

    for girl in girls_workbook:
        girls[girl.title] = []
        for boy in boys_workbook:
            score = 0
            for i in range(2, 27):
                match i:
                    case 3:
                        if girl["B" + str(i)].value == boy["B" + str(i - 1)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i - 1)].value:
                            score += -15

                    case 4:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 10
                    case 5:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 6:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 7:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 8:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 9:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 10:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 10
                    case 11:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 12:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 13:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 14:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 15:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 16:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 17:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 18:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 19:
                        if girl["B" + str(i)].value == boy["B" + str(i+1)].value:
                            score += 10
                    case 21:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 2.5
                    case 22:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 5
                    case 23:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 2.5
                    case 24:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 10
                    case 25:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15
                    case 26:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

            if girl.title == boy.title:
                girls[girl.title].append(Item(boy.title, 0))
            elif girl.title != boy.title:
                girls[girl.title].append(Item(boy.title, score))

    for boy in boys_workbook:
        boys[boy.title] = {}
        attribution[boy.title] = None
        for girl in girls_workbook:
            score = 0
            for i in range(2, 27):
                match i:
                    case 3:
                        if boy["B" + str(i)].value == girl["B" + str(i - 1)].value:
                            score += 15
                        elif boy["B" + str(i)].value != girl["B" + str(i - 1)].value:
                            score += -15

                    case 4:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 10
                    case 5:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 6:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 7:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 8:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 9:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 10:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 10
                    case 11:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 12:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 13:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 14:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 15:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 16:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 17:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 2.5
                    case 18:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 5
                    case 19:
                        if boy["B" + str(i)].value == girl["B" + str(i + 1)].value:
                            score += 10
                    case 21:
                        if boy["B" + str(i)].value == girl["B" + str(i + 1)].value:
                            score += 2.5
                    case 22:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 5
                    case 23:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 2.5
                    case 24:
                        if girl["B" + str(i)].value == boy["B" + str(i + 1)].value:
                            score += 10
                    case 25:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15

                    case 26:
                        if girl["B" + str(i)].value == boy["B" + str(i)].value:
                            score += 15
                        elif girl["B" + str(i)].value != boy["B" + str(i)].value:
                            score += -15


            if girl.title == boy.title:
                boys[boy.title][girl.title] = 0
            elif girl.title != boy.title:
                boys[boy.title][girl.title] = score
    return {
        "boys": boys,
        "girls": girls,
        "attribution": attribution,
    }


