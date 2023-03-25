import json

from openpyxl import Workbook
import requests


def retrieve_data():
    girl_poll = Workbook()
    girls_request = requests.get('https://api.lagazettedeleon.social/cupideon/girls/list/',
                                 headers={"Authorization": "Token NOTHING TO SEE HERE"})
    girls = girls_request.json
    for girl in girls():
        id = girl['first_name'] + "." + girl['last_name'] + "." + girl['classe']
        girl_poll.create_sheet(id)
        girl_poll.active = girl_poll[id]
        sheet = girl_poll.active
        sheet['A1'] = "MAIL"
        sheet['B1'] = girl['mail']
        for i in range(2, 27):
            sheet['A' + str(i)] = "Q" + str(i - 1)
            sheet['B' + str(i)] = girl['q' + str(i - 1)]
        if sheet['B9'] == 'B':
            print("DINO")
    girl_poll.save('poll/GARCONS.xlsx')

    boy_poll = Workbook()
    boys_request = requests.get('https://api.lagazettedeleon.social/cupideon/boys/list/',
                                headers={"Authorization": "Token NOTHING TO SEE HERE"})
    boys = boys_request.json
    for boy in boys():
        id = boy['first_name'] + "." + boy['last_name'] + "." + boy['classe']
        boy_poll.create_sheet(id)
        boy_poll.active = boy_poll[id]
        sheet = boy_poll.active
        sheet['A1'] = "MAIL"
        sheet['B1'] = boy['mail']
        for i in range(2, 27):
            sheet['A' + str(i)] = "Q" + str(i - 1)
            sheet['B' + str(i)] = boy['q' + str(i - 1)]
        if sheet['B9'] == 'B':
            print("DINO")
    boy_poll.save('poll/FILLES.xlsx')


